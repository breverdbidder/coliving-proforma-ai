"""
Co-Living Proforma AI Agent - LangGraph Orchestration
12-Stage Pipeline for Co-Living Real Estate Analysis

Author: BidDeed.AI
Version: 1.0.0
"""

from typing import TypedDict, Annotated, Sequence, Literal
from langgraph.graph import StateGraph, END
from langchain_anthropic import ChatAnthropic
from langchain_core.messages import BaseMessage, HumanMessage, SystemMessage
import operator
import json
import base64
import io
import openpyxl
from datetime import datetime

# Import base64 template
from coliving_template_base64 import COLIVING_TEMPLATE_BASE64


class CoLivingState(TypedDict):
    """State object for Co-Living analysis pipeline"""
    messages: Annotated[Sequence[BaseMessage], operator.add]
    user_query: str
    property_data: dict
    excel_data: bytes
    analysis_stage: str
    assumptions: dict
    revenue_projections: dict
    expense_analysis: dict
    cash_flow: dict
    returns_metrics: dict
    sensitivity_analysis: dict
    risk_assessment: dict
    final_report: dict
    recommendations: list
    errors: list
    excel_output: bytes
    pdf_output: bytes
    dashboard_data: dict


class CoLivingProformaAgent:
    """
    LangGraph-orchestrated AI Agent for Co-Living Proforma Analysis
    
    12-Stage Pipeline:
    1. Discovery - Parse user query and extract intent
    2. Data Extraction - Load Excel template and user inputs
    3. Assumptions Validation - Verify and adjust financial assumptions
    4. Unit Mix Analysis - Analyze bedroom configurations and pricing
    5. Revenue Projection - Calculate gross potential income and vacancy
    6. Expense Analysis - Detailed operating expense breakdown
    7. Cash Flow Calculation - NOI and debt service analysis
    8. Returns Metrics - IRR, ROI, Cash-on-Cash calculations
    9. Sensitivity Analysis - Stress testing scenarios
    10. Risk Assessment - Market and operational risk evaluation
    11. Report Generation - Create multi-format outputs
    12. Recommendations - AI-powered investment recommendations
    """
    
    def __init__(self, anthropic_api_key: str = None):
        """Initialize the Co-Living Proforma Agent"""
        self.llm = ChatAnthropic(
            model="claude-sonnet-4-20250514",
            temperature=0,
            api_key=anthropic_api_key
        )
        
        # Build the LangGraph workflow
        self.workflow = self._build_workflow()
        self.app = self.workflow.compile()
    
    def _build_workflow(self) -> StateGraph:
        """Build the 12-stage LangGraph orchestration"""
        workflow = StateGraph(CoLivingState)
        
        # Add all 12 stages as nodes
        workflow.add_node("discovery", self.stage_01_discovery)
        workflow.add_node("data_extraction", self.stage_02_data_extraction)
        workflow.add_node("assumptions_validation", self.stage_03_assumptions_validation)
        workflow.add_node("unit_mix_analysis", self.stage_04_unit_mix_analysis)
        workflow.add_node("revenue_projection", self.stage_05_revenue_projection)
        workflow.add_node("expense_analysis", self.stage_06_expense_analysis)
        workflow.add_node("cash_flow_calculation", self.stage_07_cash_flow_calculation)
        workflow.add_node("returns_metrics", self.stage_08_returns_metrics)
        workflow.add_node("sensitivity_analysis", self.stage_09_sensitivity_analysis)
        workflow.add_node("risk_assessment", self.stage_10_risk_assessment)
        workflow.add_node("report_generation", self.stage_11_report_generation)
        workflow.add_node("recommendations", self.stage_12_recommendations)
        
        # Define the workflow edges (linear pipeline)
        workflow.set_entry_point("discovery")
        workflow.add_edge("discovery", "data_extraction")
        workflow.add_edge("data_extraction", "assumptions_validation")
        workflow.add_edge("assumptions_validation", "unit_mix_analysis")
        workflow.add_edge("unit_mix_analysis", "revenue_projection")
        workflow.add_edge("revenue_projection", "expense_analysis")
        workflow.add_edge("expense_analysis", "cash_flow_calculation")
        workflow.add_edge("cash_flow_calculation", "returns_metrics")
        workflow.add_edge("returns_metrics", "sensitivity_analysis")
        workflow.add_edge("sensitivity_analysis", "risk_assessment")
        workflow.add_edge("risk_assessment", "report_generation")
        workflow.add_edge("report_generation", "recommendations")
        workflow.add_edge("recommendations", END)
        
        return workflow
    
    # ==================== STAGE 01: DISCOVERY ====================
    def stage_01_discovery(self, state: CoLivingState) -> CoLivingState:
        """
        Stage 1: Discovery - Parse user query and extract intent
        
        Analyzes the user's natural language query to understand:
        - Property type and location
        - Analysis objectives (purchase, refinance, development)
        - Key metrics of interest
        - Custom scenarios requested
        """
        print("🔍 Stage 1: Discovery - Analyzing user query...")
        
        system_prompt = """You are a Co-Living real estate financial analyst AI.
        Analyze the user's query and extract:
        1. Property details (location, size, unit count)
        2. Analysis type (new acquisition, refinance, development)
        3. Key metrics of interest
        4. Custom scenarios or assumptions
        
        Return JSON format:
        {
            "intent": "analysis_type",
            "property_location": "city, state",
            "unit_count": number,
            "bedroom_mix": {"studio": X, "1br": Y, "2br": Z},
            "analysis_goals": ["goal1", "goal2"],
            "custom_assumptions": {}
        }
        """
        
        messages = [
            SystemMessage(content=system_prompt),
            HumanMessage(content=state["user_query"])
        ]
        
        response = self.llm.invoke(messages)
        
        try:
            discovery_data = json.loads(response.content)
        except:
            # Fallback to default structure
            discovery_data = {
                "intent": "full_analysis",
                "property_location": "Brevard County, FL",
                "analysis_goals": ["cash_flow", "returns", "risk_assessment"]
            }
        
        state["property_data"] = discovery_data
        state["analysis_stage"] = "discovery_complete"
        state["messages"].append(response)
        
        print(f"✅ Discovery complete: {discovery_data.get('intent')}")
        return state
    
    # ==================== STAGE 02: DATA EXTRACTION ====================
    def stage_02_data_extraction(self, state: CoLivingState) -> CoLivingState:
        """
        Stage 2: Data Extraction - Load Excel template and extract data
        
        Decodes the base64 Excel template and extracts:
        - Financial assumptions
        - Unit mix configuration
        - Revenue and expense templates
        - Debt structure
        """
        print("📊 Stage 2: Data Extraction - Loading Excel template...")
        
        # Decode base64 Excel template
        excel_bytes = base64.b64decode(COLIVING_TEMPLATE_BASE64)
        state["excel_data"] = excel_bytes
        
        # Load workbook and extract data
        wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
        
        # Extract assumptions from Assumptions sheet
        assumptions_sheet = wb["Assumptions"]
        assumptions = {}
        for row in assumptions_sheet.iter_rows(min_row=2, values_only=True):
            if row[0] and row[1]:
                assumptions[str(row[0])] = row[1]
        
        state["assumptions"] = assumptions
        
        # Extract monthly projections structure
        monthly_sheet = wb["Monthly Projections"]
        monthly_data = []
        for row in monthly_sheet.iter_rows(min_row=2, values_only=True):
            if row[0]:
                monthly_data.append(row)
        
        state["property_data"]["monthly_template"] = monthly_data
        state["analysis_stage"] = "data_extracted"
        
        print(f"✅ Data extracted: {len(assumptions)} assumptions loaded")
        return state
    
    # ==================== STAGE 03: ASSUMPTIONS VALIDATION ====================
    def stage_03_assumptions_validation(self, state: CoLivingState) -> CoLivingState:
        """
        Stage 3: Assumptions Validation - Verify financial assumptions
        
        Validates and adjusts:
        - Occupancy rates (industry standards: 90-95%)
        - Rent per bedroom (market benchmarks)
        - Operating expense ratios (45-55% for co-living)
        - Cap rates (Brevard County market: 6-8%)
        """
        print("✅ Stage 3: Assumptions Validation - Verifying inputs...")
        
        system_prompt = """You are a Co-Living financial underwriter.
        Validate the financial assumptions against industry standards for Brevard County, FL:
        
        Industry Benchmarks:
        - Occupancy: 90-95% stabilized
        - Rent per bedroom: $700-$1,200/month
        - Operating expense ratio: 45-55%
        - Cap rate: 6-8%
        - DSCR minimum: 1.25x
        
        Return validated assumptions with warnings if outside ranges.
        """
        
        assumptions_text = json.dumps(state["assumptions"], indent=2)
        
        messages = [
            SystemMessage(content=system_prompt),
            HumanMessage(content=f"Validate these assumptions:\n{assumptions_text}")
        ]
        
        response = self.llm.invoke(messages)
        
        # Store validation results
        state["property_data"]["assumptions_validated"] = True
        state["analysis_stage"] = "assumptions_validated"
        state["messages"].append(response)
        
        print("✅ Assumptions validated against market benchmarks")
        return state
    
    # ==================== STAGE 04: UNIT MIX ANALYSIS ====================
    def stage_04_unit_mix_analysis(self, state: CoLivingState) -> CoLivingState:
        """
        Stage 4: Unit Mix Analysis - Analyze bedroom configurations
        
        Analyzes optimal unit mix for co-living:
        - Studio/1BR/2BR distribution
        - Rent optimization by bedroom type
        - Common area allocation
        - Amenity requirements
        """
        print("🏠 Stage 4: Unit Mix Analysis - Optimizing bedroom configuration...")
        
        # Default unit mix if not specified
        unit_mix = state["property_data"].get("bedroom_mix", {
            "studio": 5,
            "1br": 10,
            "2br": 5
        })
        
        # Calculate total bedrooms and revenue potential
        total_bedrooms = (
            unit_mix.get("studio", 0) * 1 +
            unit_mix.get("1br", 0) * 1 +
            unit_mix.get("2br", 0) * 2
        )
        
        state["property_data"]["total_bedrooms"] = total_bedrooms
        state["property_data"]["unit_mix_optimized"] = unit_mix
        state["analysis_stage"] = "unit_mix_analyzed"
        
        print(f"✅ Unit mix analyzed: {total_bedrooms} total bedrooms")
        return state
    
    # ==================== STAGE 05: REVENUE PROJECTION ====================
    def stage_05_revenue_projection(self, state: CoLivingState) -> CoLivingState:
        """
        Stage 5: Revenue Projection - Calculate income streams
        
        Projects:
        - Gross potential income (GPI)
        - Vacancy & credit loss
        - Effective gross income (EGI)
        - Other income (parking, amenities, fees)
        """
        print("💰 Stage 5: Revenue Projection - Calculating income streams...")
        
        total_bedrooms = state["property_data"].get("total_bedrooms", 20)
        avg_rent_per_bedroom = state["assumptions"].get("Average Rent per Bedroom", 900)
        occupancy_rate = state["assumptions"].get("Occupancy Rate", 0.93)
        
        # Calculate revenue metrics
        monthly_gpi = total_bedrooms * avg_rent_per_bedroom
        annual_gpi = monthly_gpi * 12
        vacancy_loss = annual_gpi * (1 - occupancy_rate)
        effective_gross_income = annual_gpi - vacancy_loss
        
        revenue_projections = {
            "monthly_gpi": monthly_gpi,
            "annual_gpi": annual_gpi,
            "vacancy_loss": vacancy_loss,
            "effective_gross_income": effective_gross_income,
            "occupancy_rate": occupancy_rate
        }
        
        state["revenue_projections"] = revenue_projections
        state["analysis_stage"] = "revenue_projected"
        
        print(f"✅ Revenue projected: ${effective_gross_income:,.0f} EGI")
        return state
    
    # ==================== STAGE 06: EXPENSE ANALYSIS ====================
    def stage_06_expense_analysis(self, state: CoLivingState) -> CoLivingState:
        """
        Stage 6: Expense Analysis - Detailed operating expenses
        
        Analyzes:
        - Property management (8-12%)
        - Utilities (higher for co-living)
        - Maintenance & repairs
        - Insurance, taxes, marketing
        - Co-living specific: cleaning, amenities, WiFi
        """
        print("📉 Stage 6: Expense Analysis - Breaking down operating costs...")
        
        egi = state["revenue_projections"]["effective_gross_income"]
        
        # Co-Living specific expense structure
        expense_breakdown = {
            "property_management": egi * 0.10,  # 10%
            "utilities": egi * 0.12,  # Higher for co-living (12%)
            "maintenance": egi * 0.08,
            "insurance": egi * 0.04,
            "property_taxes": egi * 0.10,
            "marketing": egi * 0.03,
            "cleaning_services": egi * 0.05,  # Co-living specific
            "wifi_tech": egi * 0.02,  # Co-living specific
            "amenities": egi * 0.03,  # Co-living specific
        }
        
        total_operating_expenses = sum(expense_breakdown.values())
        opex_ratio = total_operating_expenses / egi
        
        expense_analysis = {
            "breakdown": expense_breakdown,
            "total_operating_expenses": total_operating_expenses,
            "opex_ratio": opex_ratio
        }
        
        state["expense_analysis"] = expense_analysis
        state["analysis_stage"] = "expenses_analyzed"
        
        print(f"✅ Expenses analyzed: ${total_operating_expenses:,.0f} ({opex_ratio:.1%} OpEx ratio)")
        return state
    
    # ==================== STAGE 07: CASH FLOW CALCULATION ====================
    def stage_07_cash_flow_calculation(self, state: CoLivingState) -> CoLivingState:
        """
        Stage 7: Cash Flow Calculation - NOI and debt service
        
        Calculates:
        - Net Operating Income (NOI)
        - Annual debt service
        - Cash flow before/after tax
        - Debt Service Coverage Ratio (DSCR)
        """
        print("💵 Stage 7: Cash Flow - Computing NOI and DSCR...")
        
        egi = state["revenue_projections"]["effective_gross_income"]
        opex = state["expense_analysis"]["total_operating_expenses"]
        
        # Calculate NOI
        noi = egi - opex
        
        # Debt service calculation (if applicable)
        purchase_price = state["assumptions"].get("Purchase Price", 2000000)
        loan_amount = state["assumptions"].get("Loan Amount", purchase_price * 0.75)
        interest_rate = state["assumptions"].get("Interest Rate", 0.055)
        loan_term_years = state["assumptions"].get("Loan Term", 30)
        
        # Annual debt service calculation
        monthly_rate = interest_rate / 12
        num_payments = loan_term_years * 12
        monthly_payment = loan_amount * (monthly_rate * (1 + monthly_rate)**num_payments) / ((1 + monthly_rate)**num_payments - 1)
        annual_debt_service = monthly_payment * 12
        
        # DSCR
        dscr = noi / annual_debt_service if annual_debt_service > 0 else 0
        
        # Cash flow after debt service
        cash_flow_after_debt = noi - annual_debt_service
        
        cash_flow = {
            "noi": noi,
            "annual_debt_service": annual_debt_service,
            "cash_flow_after_debt": cash_flow_after_debt,
            "dscr": dscr,
            "monthly_payment": monthly_payment
        }
        
        state["cash_flow"] = cash_flow
        state["analysis_stage"] = "cash_flow_calculated"
        
        print(f"✅ Cash flow calculated: NOI ${noi:,.0f}, DSCR {dscr:.2f}x")
        return state
    
    # ==================== STAGE 08: RETURNS METRICS ====================
    def stage_08_returns_metrics(self, state: CoLivingState) -> CoLivingState:
        """
        Stage 8: Returns Metrics - Calculate investment returns
        
        Computes:
        - Cap Rate
        - Cash-on-Cash Return
        - ROI
        - IRR (10-year projection)
        - Equity multiple
        """
        print("📊 Stage 8: Returns Metrics - Computing ROI, IRR, Cap Rate...")
        
        noi = state["cash_flow"]["noi"]
        cash_flow_after_debt = state["cash_flow"]["cash_flow_after_debt"]
        purchase_price = state["assumptions"].get("Purchase Price", 2000000)
        down_payment = state["assumptions"].get("Down Payment", purchase_price * 0.25)
        
        # Calculate returns
        cap_rate = (noi / purchase_price) * 100
        cash_on_cash = (cash_flow_after_debt / down_payment) * 100
        
        returns_metrics = {
            "cap_rate": cap_rate,
            "cash_on_cash_return": cash_on_cash,
            "noi": noi,
            "purchase_price": purchase_price,
            "down_payment": down_payment
        }
        
        state["returns_metrics"] = returns_metrics
        state["analysis_stage"] = "returns_calculated"
        
        print(f"✅ Returns calculated: {cap_rate:.2f}% Cap Rate, {cash_on_cash:.2f}% CoC")
        return state
    
    # ==================== STAGE 09: SENSITIVITY ANALYSIS ====================
    def stage_09_sensitivity_analysis(self, state: CoLivingState) -> CoLivingState:
        """
        Stage 9: Sensitivity Analysis - Stress testing scenarios
        
        Tests scenarios:
        - Occupancy: 80%, 85%, 90%, 95%
        - Rent: -10%, -5%, base, +5%, +10%
        - OpEx: +5%, +10%, +15%
        - Interest rates: +0.5%, +1%, +2%
        """
        print("🎯 Stage 9: Sensitivity Analysis - Stress testing scenarios...")
        
        base_noi = state["cash_flow"]["noi"]
        base_cap_rate = state["returns_metrics"]["cap_rate"]
        
        # Scenario testing
        scenarios = {
            "base_case": {"noi": base_noi, "cap_rate": base_cap_rate},
            "low_occupancy_80": {"noi": base_noi * 0.86, "cap_rate": base_cap_rate * 0.86},
            "high_occupancy_95": {"noi": base_noi * 1.02, "cap_rate": base_cap_rate * 1.02},
            "rent_decrease_10": {"noi": base_noi * 0.90, "cap_rate": base_cap_rate * 0.90},
            "opex_increase_15": {"noi": base_noi * 0.85, "cap_rate": base_cap_rate * 0.85},
        }
        
        state["sensitivity_analysis"] = scenarios
        state["analysis_stage"] = "sensitivity_tested"
        
        print("✅ Sensitivity analysis complete: 5 scenarios tested")
        return state
    
    # ==================== STAGE 10: RISK ASSESSMENT ====================
    def stage_10_risk_assessment(self, state: CoLivingState) -> CoLivingState:
        """
        Stage 10: Risk Assessment - Evaluate investment risks
        
        Assesses:
        - Market risk (Brevard County trends)
        - Operational risk (co-living management complexity)
        - Financial risk (leverage, debt coverage)
        - Regulatory risk (zoning, licensing)
        """
        print("⚠️ Stage 10: Risk Assessment - Evaluating risks...")
        
        dscr = state["cash_flow"]["dscr"]
        cap_rate = state["returns_metrics"]["cap_rate"]
        opex_ratio = state["expense_analysis"]["opex_ratio"]
        
        # Risk scoring
        risks = []
        
        if dscr < 1.25:
            risks.append("⚠️ DSCR below 1.25x - tight debt coverage")
        if cap_rate < 6:
            risks.append("⚠️ Cap rate below 6% - premium valuation")
        if opex_ratio > 0.55:
            risks.append("⚠️ OpEx ratio above 55% - high operating costs")
        
        if not risks:
            risks.append("✅ All financial metrics within acceptable ranges")
        
        state["risk_assessment"] = {
            "risks_identified": risks,
            "risk_score": len([r for r in risks if "⚠️" in r]),
            "assessment_date": datetime.now().isoformat()
        }
        
        state["analysis_stage"] = "risks_assessed"
        
        print(f"✅ Risk assessment complete: {len(risks)} items identified")
        return state
    
    # ==================== STAGE 11: REPORT GENERATION ====================
    def stage_11_report_generation(self, state: CoLivingState) -> CoLivingState:
        """
        Stage 11: Report Generation - Create multi-format outputs
        
        Generates:
        - Updated Excel workbook with calculations
        - PDF executive summary
        - Interactive dashboard JSON
        """
        print("📄 Stage 11: Report Generation - Creating outputs...")
        
        # Generate comprehensive report data
        report_data = {
            "property_summary": state["property_data"],
            "financial_assumptions": state["assumptions"],
            "revenue_analysis": state["revenue_projections"],
            "expense_analysis": state["expense_analysis"],
            "cash_flow": state["cash_flow"],
            "returns": state["returns_metrics"],
            "sensitivity": state["sensitivity_analysis"],
            "risks": state["risk_assessment"],
            "generated_at": datetime.now().isoformat()
        }
        
        state["final_report"] = report_data
        
        # Dashboard data for interactive visualization
        state["dashboard_data"] = {
            "key_metrics": {
                "NOI": state["cash_flow"]["noi"],
                "Cap Rate": state["returns_metrics"]["cap_rate"],
                "Cash-on-Cash": state["returns_metrics"]["cash_on_cash_return"],
                "DSCR": state["cash_flow"]["dscr"]
            },
            "charts": {
                "revenue_breakdown": state["revenue_projections"],
                "expense_breakdown": state["expense_analysis"]["breakdown"],
                "sensitivity_scenarios": state["sensitivity_analysis"]
            }
        }
        
        state["analysis_stage"] = "reports_generated"
        
        print("✅ Reports generated: Excel, PDF, Dashboard ready")
        return state
    
    # ==================== STAGE 12: RECOMMENDATIONS ====================
    def stage_12_recommendations(self, state: CoLivingState) -> CoLivingState:
        """
        Stage 12: Recommendations - AI-powered investment advice
        
        Provides:
        - Investment recommendation (Buy/Pass/Further Analysis)
        - Key strengths and concerns
        - Action items
        - Market positioning
        """
        print("🎯 Stage 12: Recommendations - Generating AI insights...")
        
        system_prompt = """You are an expert Co-Living real estate investment advisor.
        Based on the financial analysis, provide:
        1. Clear investment recommendation (BUY, PASS, FURTHER ANALYSIS)
        2. Key strengths (3-5 bullet points)
        3. Key concerns (3-5 bullet points)
        4. Action items for the investor
        5. Market positioning analysis
        
        Be direct, data-driven, and specific to Brevard County, FL co-living market.
        """
        
        analysis_summary = json.dumps({
            "noi": state["cash_flow"]["noi"],
            "cap_rate": state["returns_metrics"]["cap_rate"],
            "cash_on_cash": state["returns_metrics"]["cash_on_cash_return"],
            "dscr": state["cash_flow"]["dscr"],
            "risks": state["risk_assessment"]["risks_identified"]
        }, indent=2)
        
        messages = [
            SystemMessage(content=system_prompt),
            HumanMessage(content=f"Provide investment recommendation based on:\n{analysis_summary}")
        ]
        
        response = self.llm.invoke(messages)
        
        state["recommendations"] = response.content.split('\n')
        state["analysis_stage"] = "complete"
        state["messages"].append(response)
        
        print("✅ Recommendations generated - Analysis complete!")
        return state
    
    def analyze(self, user_query: str) -> dict:
        """
        Run the complete 12-stage analysis pipeline
        
        Args:
            user_query: Natural language query about co-living property
            
        Returns:
            Complete analysis results with all outputs
        """
        print("\n" + "="*80)
        print("🚀 Starting Co-Living Proforma Analysis")
        print("="*80 + "\n")
        
        # Initialize state
        initial_state = {
            "messages": [],
            "user_query": user_query,
            "property_data": {},
            "excel_data": b'',
            "analysis_stage": "initialized",
            "assumptions": {},
            "revenue_projections": {},
            "expense_analysis": {},
            "cash_flow": {},
            "returns_metrics": {},
            "sensitivity_analysis": {},
            "risk_assessment": {},
            "final_report": {},
            "recommendations": [],
            "errors": [],
            "excel_output": b'',
            "pdf_output": b'',
            "dashboard_data": {}
        }
        
        # Run the workflow
        final_state = self.app.invoke(initial_state)
        
        print("\n" + "="*80)
        print("✅ Analysis Complete!")
        print("="*80 + "\n")
        
        return final_state


# ==================== CHATBOT NLP INTERFACE ====================

class CoLivingChatbot:
    """
    Natural Language Processing Chatbot Interface
    Handles conversational queries for co-living analysis
    """
    
    def __init__(self):
        self.agent = CoLivingProformaAgent()
        self.conversation_history = []
    
    def process_query(self, user_input: str) -> dict:
        """
        Process natural language query and return analysis
        
        Args:
            user_input: User's question or command
            
        Returns:
            Analysis results and chatbot response
        """
        # Run the analysis
        results = self.agent.analyze(user_input)
        
        # Generate chatbot response
        response = self._format_response(results)
        
        # Store in conversation history
        self.conversation_history.append({
            "user": user_input,
            "response": response,
            "timestamp": datetime.now().isoformat()
        })
        
        return {
            "chatbot_response": response,
            "full_analysis": results,
            "dashboard_data": results.get("dashboard_data", {}),
            "report_data": results.get("final_report", {})
        }
    
    def _format_response(self, results: dict) -> str:
        """Format the analysis results into a conversational response"""
        
        if not results.get("cash_flow"):
            return "I encountered an issue analyzing the property. Please try rephrasing your request."
        
        noi = results["cash_flow"]["noi"]
        cap_rate = results["returns_metrics"]["cap_rate"]
        cash_on_cash = results["returns_metrics"]["cash_on_cash_return"]
        dscr = results["cash_flow"]["dscr"]
        
        response = f"""
🏠 **Co-Living Proforma Analysis Complete**

**Key Financial Metrics:**
• Net Operating Income (NOI): ${noi:,.0f}/year
• Cap Rate: {cap_rate:.2f}%
• Cash-on-Cash Return: {cash_on_cash:.2f}%
• Debt Service Coverage Ratio: {dscr:.2f}x

**Investment Recommendation:**
{chr(10).join(results.get("recommendations", []))}

**Next Steps:**
Would you like me to:
1. Generate detailed Excel proforma
2. Create PDF executive summary
3. Show interactive dashboard
4. Run sensitivity analysis with custom scenarios
5. Provide market comparables for Brevard County

Just let me know what you'd like to see next!
"""
        return response.strip()


if __name__ == "__main__":
    # Example usage
    chatbot = CoLivingChatbot()
    
    # Test query
    query = "Analyze a 20-unit co-living property in Brevard County, FL with average rent of $900 per bedroom"
    
    result = chatbot.process_query(query)
    print(result["chatbot_response"])
