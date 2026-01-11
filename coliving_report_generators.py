"""
Co-Living Proforma Report Generators
Generates Excel, PDF, and Dashboard outputs

Author: BidDeed.AI
Version: 1.0.0
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import LineChart, BarChart, Reference
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
import io
import json
from datetime import datetime
import base64


class CoLivingReportGenerator:
    """
    Multi-format report generation for Co-Living analysis
    Outputs: Excel, PDF, Dashboard JSON
    """
    
    def __init__(self, analysis_results: dict):
        """
        Initialize report generator with analysis results
        
        Args:
            analysis_results: Complete output from CoLivingProformaAgent
        """
        self.results = analysis_results
        self.property_data = analysis_results.get("property_data", {})
        self.assumptions = analysis_results.get("assumptions", {})
        self.revenue = analysis_results.get("revenue_projections", {})
        self.expenses = analysis_results.get("expense_analysis", {})
        self.cash_flow = analysis_results.get("cash_flow", {})
        self.returns = analysis_results.get("returns_metrics", {})
        self.sensitivity = analysis_results.get("sensitivity_analysis", {})
        self.risks = analysis_results.get("risk_assessment", {})
        self.recommendations = analysis_results.get("recommendations", [])
    
    # ==================== EXCEL REPORT GENERATOR ====================
    
    def generate_excel_report(self) -> bytes:
        """
        Generate comprehensive Excel proforma workbook
        
        Returns:
            Bytes of Excel file
        """
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create all sheets
        self._create_executive_summary_sheet(wb)
        self._create_assumptions_sheet(wb)
        self._create_revenue_analysis_sheet(wb)
        self._create_expense_analysis_sheet(wb)
        self._create_cash_flow_sheet(wb)
        self._create_returns_analysis_sheet(wb)
        self._create_sensitivity_sheet(wb)
        self._create_risk_assessment_sheet(wb)
        self._create_charts_dashboard_sheet(wb)
        
        # Save to bytes
        excel_bytes = io.BytesIO()
        wb.save(excel_bytes)
        excel_bytes.seek(0)
        
        return excel_bytes.getvalue()
    
    def _create_executive_summary_sheet(self, wb):
        """Create Executive Summary sheet"""
        ws = wb.create_sheet("Executive Summary", 0)
        
        # Header styling
        header_fill = PatternFill(start_color="2563eb", end_color="2563eb", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=14)
        
        # Title
        ws['A1'] = "BidDeed.AI Co-Living Proforma"
        ws['A1'].font = Font(bold=True, size=18, color="2563eb")
        ws.merge_cells('A1:C1')
        
        ws['A2'] = f"Generated: {datetime.now().strftime('%B %d, %Y')}"
        ws['A2'].font = Font(size=10, color="6b7280")
        
        # Property Info
        row = 4
        ws[f'A{row}'] = "PROPERTY INFORMATION"
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        ws.merge_cells(f'A{row}:C{row}')
        
        row += 1
        property_info = [
            ("Location", self.property_data.get("property_location", "Brevard County, FL")),
            ("Total Bedrooms", self.property_data.get("total_bedrooms", "N/A")),
            ("Analysis Type", self.property_data.get("intent", "Full Analysis"))
        ]
        
        for label, value in property_info:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        # Key Metrics
        row += 1
        ws[f'A{row}'] = "KEY FINANCIAL METRICS"
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        ws.merge_cells(f'A{row}:C{row}')
        
        row += 1
        metrics = [
            ("Net Operating Income (NOI)", f"${self.cash_flow.get('noi', 0):,.0f}"),
            ("Cap Rate", f"{self.returns.get('cap_rate', 0):.2f}%"),
            ("Cash-on-Cash Return", f"{self.returns.get('cash_on_cash_return', 0):.2f}%"),
            ("DSCR", f"{self.cash_flow.get('dscr', 0):.2f}x"),
            ("Purchase Price", f"${self.returns.get('purchase_price', 0):,.0f}")
        ]
        
        for label, value in metrics:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'].font = Font(size=12, color="10b981")
            row += 1
        
        # Investment Recommendation
        row += 1
        ws[f'A{row}'] = "INVESTMENT RECOMMENDATION"
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        ws.merge_cells(f'A{row}:C{row}')
        
        row += 1
        for rec in self.recommendations[:5]:
            ws[f'A{row}'] = rec
            ws[f'A{row}'].alignment = Alignment(wrap_text=True)
            row += 1
        
        # Column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
    
    def _create_assumptions_sheet(self, wb):
        """Create Assumptions sheet"""
        ws = wb.create_sheet("Assumptions")
        
        ws['A1'] = "Financial Assumptions"
        ws['A1'].font = Font(bold=True, size=14)
        
        row = 3
        ws[f'A{row}'] = "Assumption"
        ws[f'B{row}'] = "Value"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].font = Font(bold=True)
        
        row += 1
        for key, value in self.assumptions.items():
            ws[f'A{row}'] = key
            ws[f'B{row}'] = value
            row += 1
        
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 20
    
    def _create_revenue_analysis_sheet(self, wb):
        """Create Revenue Analysis sheet"""
        ws = wb.create_sheet("Revenue Analysis")
        
        ws['A1'] = "Revenue Projections"
        ws['A1'].font = Font(bold=True, size=14)
        
        row = 3
        revenue_items = [
            ("Metric", "Annual Amount"),
            ("Gross Potential Income", f"${self.revenue.get('annual_gpi', 0):,.0f}"),
            ("Vacancy Loss", f"${self.revenue.get('vacancy_loss', 0):,.0f}"),
            ("Effective Gross Income", f"${self.revenue.get('effective_gross_income', 0):,.0f}"),
            ("Occupancy Rate", f"{self.revenue.get('occupancy_rate', 0):.1%}")
        ]
        
        for label, value in revenue_items:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            if row > 3:
                ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
    
    def _create_expense_analysis_sheet(self, wb):
        """Create Expense Analysis sheet"""
        ws = wb.create_sheet("Expense Analysis")
        
        ws['A1'] = "Operating Expense Breakdown"
        ws['A1'].font = Font(bold=True, size=14)
        
        row = 3
        ws[f'A{row}'] = "Expense Category"
        ws[f'B{row}'] = "Annual Amount"
        ws[f'C{row}'] = "% of EGI"
        
        for col in ['A', 'B', 'C']:
            ws[f'{col}{row}'].font = Font(bold=True)
        
        row += 1
        egi = self.revenue.get('effective_gross_income', 1)
        
        for category, amount in self.expenses.get('breakdown', {}).items():
            ws[f'A{row}'] = category.replace('_', ' ').title()
            ws[f'B{row}'] = f"${amount:,.0f}"
            ws[f'C{row}'] = f"{(amount/egi)*100:.1f}%"
            row += 1
        
        row += 1
        ws[f'A{row}'] = "Total Operating Expenses"
        ws[f'B{row}'] = f"${self.expenses.get('total_operating_expenses', 0):,.0f}"
        ws[f'C{row}'] = f"{self.expenses.get('opex_ratio', 0):.1%}"
        
        for col in ['A', 'B', 'C']:
            ws[f'{col}{row}'].font = Font(bold=True, color="ef4444")
        
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
    
    def _create_cash_flow_sheet(self, wb):
        """Create Cash Flow sheet"""
        ws = wb.create_sheet("Cash Flow Analysis")
        
        ws['A1'] = "Cash Flow Statement"
        ws['A1'].font = Font(bold=True, size=14)
        
        row = 3
        cash_flow_items = [
            ("Effective Gross Income", f"${self.revenue.get('effective_gross_income', 0):,.0f}"),
            ("- Total Operating Expenses", f"${self.expenses.get('total_operating_expenses', 0):,.0f}"),
            ("= Net Operating Income (NOI)", f"${self.cash_flow.get('noi', 0):,.0f}"),
            ("- Annual Debt Service", f"${self.cash_flow.get('annual_debt_service', 0):,.0f}"),
            ("= Cash Flow After Debt", f"${self.cash_flow.get('cash_flow_after_debt', 0):,.0f}"),
            ("", ""),
            ("Debt Service Coverage Ratio", f"{self.cash_flow.get('dscr', 0):.2f}x")
        ]
        
        for label, value in cash_flow_items:
            if label:
                ws[f'A{row}'] = label
                ws[f'B{row}'] = value
                if "=" in label:
                    ws[f'A{row}'].font = Font(bold=True, color="10b981")
                    ws[f'B{row}'].font = Font(bold=True, color="10b981")
            row += 1
        
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 20
    
    def _create_returns_analysis_sheet(self, wb):
        """Create Returns Analysis sheet"""
        ws = wb.create_sheet("Returns Analysis")
        
        ws['A1'] = "Investment Returns Metrics"
        ws['A1'].font = Font(bold=True, size=14)
        
        row = 3
        returns_items = [
            ("Cap Rate", f"{self.returns.get('cap_rate', 0):.2f}%"),
            ("Cash-on-Cash Return", f"{self.returns.get('cash_on_cash_return', 0):.2f}%"),
            ("Purchase Price", f"${self.returns.get('purchase_price', 0):,.0f}"),
            ("Down Payment (25%)", f"${self.returns.get('down_payment', 0):,.0f}"),
            ("NOI", f"${self.returns.get('noi', 0):,.0f}")
        ]
        
        for label, value in returns_items:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
    
    def _create_sensitivity_sheet(self, wb):
        """Create Sensitivity Analysis sheet"""
        ws = wb.create_sheet("Sensitivity Analysis")
        
        ws['A1'] = "Sensitivity Analysis - Scenario Testing"
        ws['A1'].font = Font(bold=True, size=14)
        
        row = 3
        ws[f'A{row}'] = "Scenario"
        ws[f'B{row}'] = "NOI"
        ws[f'C{row}'] = "Cap Rate"
        
        for col in ['A', 'B', 'C']:
            ws[f'{col}{row}'].font = Font(bold=True)
        
        row += 1
        for scenario_name, scenario_data in self.sensitivity.items():
            ws[f'A{row}'] = scenario_name.replace('_', ' ').title()
            ws[f'B{row}'] = f"${scenario_data.get('noi', 0):,.0f}"
            ws[f'C{row}'] = f"{scenario_data.get('cap_rate', 0):.2f}%"
            row += 1
        
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
    
    def _create_risk_assessment_sheet(self, wb):
        """Create Risk Assessment sheet"""
        ws = wb.create_sheet("Risk Assessment")
        
        ws['A1'] = "Risk Analysis & Mitigation"
        ws['A1'].font = Font(bold=True, size=14)
        
        row = 3
        ws[f'A{row}'] = "Risk Factor"
        ws[f'A{row}'].font = Font(bold=True)
        
        row += 1
        for risk in self.risks.get('risks_identified', []):
            ws[f'A{row}'] = risk
            ws[f'A{row}'].alignment = Alignment(wrap_text=True)
            row += 1
        
        ws.column_dimensions['A'].width = 80
    
    def _create_charts_dashboard_sheet(self, wb):
        """Create Charts & Dashboard sheet"""
        ws = wb.create_sheet("Charts & Dashboard")
        
        ws['A1'] = "Visual Dashboard"
        ws['A1'].font = Font(bold=True, size=14)
        
        ws['A3'] = "Interactive charts and visualizations would be rendered here"
        ws['A3'].alignment = Alignment(wrap_text=True)
        
        ws.column_dimensions['A'].width = 60
    
    # ==================== PDF REPORT GENERATOR ====================
    
    def generate_pdf_report(self) -> bytes:
        """
        Generate executive summary PDF report
        
        Returns:
            Bytes of PDF file
        """
        pdf_buffer = io.BytesIO()
        doc = SimpleDocTemplate(pdf_buffer, pagesize=letter)
        story = []
        styles = getSampleStyleSheet()
        
        # Custom styles
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#2563eb'),
            spaceAfter=30
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=16,
            textColor=colors.HexColor('#111827'),
            spaceAfter=12
        )
        
        # Title
        story.append(Paragraph("BidDeed.AI Co-Living Proforma", title_style))
        story.append(Paragraph(f"Executive Summary - {datetime.now().strftime('%B %d, %Y')}", styles['Normal']))
        story.append(Spacer(1, 0.3*inch))
        
        # Property Information
        story.append(Paragraph("Property Information", heading_style))
        property_data = [
            ['Location:', self.property_data.get("property_location", "Brevard County, FL")],
            ['Total Bedrooms:', str(self.property_data.get("total_bedrooms", "N/A"))],
            ['Analysis Type:', self.property_data.get("intent", "Full Analysis")]
        ]
        property_table = Table(property_data, colWidths=[2*inch, 4*inch])
        property_table.setStyle(TableStyle([
            ('FONT', (0, 0), (-1, -1), 'Helvetica', 10),
            ('FONT', (0, 0), (0, -1), 'Helvetica-Bold', 10),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))
        story.append(property_table)
        story.append(Spacer(1, 0.3*inch))
        
        # Key Financial Metrics
        story.append(Paragraph("Key Financial Metrics", heading_style))
        metrics_data = [
            ['Metric', 'Value'],
            ['Net Operating Income (NOI)', f"${self.cash_flow.get('noi', 0):,.0f}"],
            ['Cap Rate', f"{self.returns.get('cap_rate', 0):.2f}%"],
            ['Cash-on-Cash Return', f"{self.returns.get('cash_on_cash_return', 0):.2f}%"],
            ['DSCR', f"{self.cash_flow.get('dscr', 0):.2f}x"],
            ['Purchase Price', f"${self.returns.get('purchase_price', 0):,.0f}"]
        ]
        metrics_table = Table(metrics_data, colWidths=[3*inch, 2*inch])
        metrics_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2563eb')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONT', (0, 0), (-1, 0), 'Helvetica-Bold', 12),
            ('FONT', (0, 1), (-1, -1), 'Helvetica', 10),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')])
        ]))
        story.append(metrics_table)
        story.append(Spacer(1, 0.3*inch))
        
        # Investment Recommendation
        story.append(Paragraph("Investment Recommendation", heading_style))
        for rec in self.recommendations[:5]:
            story.append(Paragraph(f"• {rec}", styles['Normal']))
        story.append(Spacer(1, 0.3*inch))
        
        # Risk Factors
        story.append(Paragraph("Risk Assessment", heading_style))
        for risk in self.risks.get('risks_identified', [])[:5]:
            story.append(Paragraph(f"• {risk}", styles['Normal']))
        
        # Build PDF
        doc.build(story)
        pdf_buffer.seek(0)
        
        return pdf_buffer.getvalue()
    
    # ==================== DASHBOARD JSON GENERATOR ====================
    
    def generate_dashboard_json(self) -> str:
        """
        Generate interactive dashboard data in JSON format
        
        Returns:
            JSON string with dashboard configuration
        """
        dashboard_config = {
            "metadata": {
                "title": "BidDeed.AI Co-Living Investment Dashboard",
                "generated_at": datetime.now().isoformat(),
                "property_location": self.property_data.get("property_location", "Brevard County, FL")
            },
            "key_metrics": {
                "noi": {
                    "label": "Net Operating Income",
                    "value": self.cash_flow.get('noi', 0),
                    "format": "currency",
                    "trend": "up",
                    "color": "#10b981"
                },
                "cap_rate": {
                    "label": "Cap Rate",
                    "value": self.returns.get('cap_rate', 0),
                    "format": "percentage",
                    "trend": "up",
                    "color": "#2563eb"
                },
                "cash_on_cash": {
                    "label": "Cash-on-Cash Return",
                    "value": self.returns.get('cash_on_cash_return', 0),
                    "format": "percentage",
                    "trend": "up",
                    "color": "#f59e0b"
                },
                "dscr": {
                    "label": "DSCR",
                    "value": self.cash_flow.get('dscr', 0),
                    "format": "ratio",
                    "trend": "neutral",
                    "color": "#6b7280"
                }
            },
            "charts": {
                "revenue_breakdown": {
                    "type": "pie",
                    "title": "Revenue Sources",
                    "data": self.revenue
                },
                "expense_breakdown": {
                    "type": "bar",
                    "title": "Operating Expenses",
                    "data": self.expenses.get('breakdown', {})
                },
                "cash_flow_waterfall": {
                    "type": "waterfall",
                    "title": "Cash Flow Analysis",
                    "data": {
                        "EGI": self.revenue.get('effective_gross_income', 0),
                        "OpEx": -self.expenses.get('total_operating_expenses', 0),
                        "NOI": self.cash_flow.get('noi', 0),
                        "Debt Service": -self.cash_flow.get('annual_debt_service', 0),
                        "Cash Flow": self.cash_flow.get('cash_flow_after_debt', 0)
                    }
                },
                "sensitivity_analysis": {
                    "type": "line",
                    "title": "Sensitivity Scenarios",
                    "data": self.sensitivity
                }
            },
            "tables": {
                "assumptions": self.assumptions,
                "risks": self.risks.get('risks_identified', []),
                "recommendations": self.recommendations
            }
        }
        
        return json.dumps(dashboard_config, indent=2)


# ==================== USAGE EXAMPLE ====================

if __name__ == "__main__":
    # Example: Generate reports from analysis results
    from coliving_orchestrator import CoLivingChatbot
    
    chatbot = CoLivingChatbot()
    query = "Analyze a 20-unit co-living property in Brevard County"
    results = chatbot.process_query(query)
    
    # Initialize report generator
    generator = CoLivingReportGenerator(results['full_analysis'])
    
    # Generate all report formats
    excel_bytes = generator.generate_excel_report()
    pdf_bytes = generator.generate_pdf_report()
    dashboard_json = generator.generate_dashboard_json()
    
    # Save to files
    with open('coliving_proforma.xlsx', 'wb') as f:
        f.write(excel_bytes)
    
    with open('coliving_executive_summary.pdf', 'wb') as f:
        f.write(pdf_bytes)
    
    with open('coliving_dashboard.json', 'w') as f:
        f.write(dashboard_json)
    
    print("✅ Reports generated successfully!")
    print("  - coliving_proforma.xlsx")
    print("  - coliving_executive_summary.pdf")
    print("  - coliving_dashboard.json")
